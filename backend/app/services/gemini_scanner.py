"""Gemini Vision を使った業者見積スキャン解析サービス。"""
from __future__ import annotations

import io
import json
import logging
from datetime import date
from pathlib import Path
from typing import Any

import structlog
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.enums import ScanJobFileType
from app.models.scan import ScanJob, ScanResult, ScanResultItem

logger = structlog.get_logger(__name__)

GEMINI_MODEL = "gemini-2.5-flash"
_PDF_MAX_PAGES = 10


# ── Gemini レスポンス用スキーマ ───────────────────────────────────────────────

class _ItemExtraction(BaseModel):
    """明細1行分の抽出結果。"""

    row_no: int = 0
    item_name: str | None = None
    spec: str | None = None
    unit: str | None = None
    quantity: float | None = None
    unit_price: float | None = None
    amount: float | None = None
    confidence: float = 0.5


class _ScanExtraction(BaseModel):
    """見積書全体の抽出結果。"""

    vendor_name: str | None = None
    quoted_date: str | None = None  # YYYY-MM-DD or null
    items: list[_ItemExtraction] = []
    subtotal: float | None = None
    tax: float | None = None
    total: float | None = None
    confidence_score: float = 0.5


# ── メイン処理 ────────────────────────────────────────────────────────────────

def process_file(job: ScanJob, db: Session) -> None:
    """ファイルを Gemini で解析し ScanResult / ScanResultItem を作成する。

    処理成功後に DB への add/flush まで行う。commit は呼び出し元（Celery タスク）が担う。
    """
    from google import genai
    from google.genai import types  # noqa: F401

    client = genai.Client(api_key=settings.gemini_api_key)

    file_path = Path(job.original_file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Upload file not found: {file_path}")

    contents = _build_contents(job.file_type, file_path)

    response = client.models.generate_content(
        model=GEMINI_MODEL,
        contents=contents,
        config={
            "response_mime_type": "application/json",
        },
    )

    raw_text = response.text or "{}"
    try:
        raw_json: dict[str, Any] = json.loads(raw_text)
        extraction = _ScanExtraction.model_validate(raw_json)
    except Exception as exc:
        logger.warning("gemini_parse_error", error=str(exc), raw=raw_text[:500])
        extraction = _ScanExtraction()
        raw_json = {}

    job.gemini_model = GEMINI_MODEL
    job.gemini_response_raw = raw_json

    quoted_date: date | None = None
    if extraction.quoted_date:
        try:
            quoted_date = date.fromisoformat(extraction.quoted_date[:10])
        except ValueError:
            pass

    scan_result = ScanResult(
        scan_job_id=job.id,
        vendor_name_detected=extraction.vendor_name,
        quoted_date_detected=quoted_date,
        subtotal_detected=extraction.subtotal,
        tax_detected=extraction.tax,
        total_detected=extraction.total,
        confidence_score=extraction.confidence_score,
    )
    db.add(scan_result)
    db.flush()

    for idx, item in enumerate(extraction.items):
        row_no = item.row_no if item.row_no > 0 else idx + 1
        db.add(
            ScanResultItem(
                scan_result_id=scan_result.id,
                row_no=row_no,
                item_name=item.item_name,
                spec=item.spec,
                unit=item.unit,
                quantity=item.quantity,
                unit_price=item.unit_price,
                amount=item.amount,
                confidence=item.confidence,
            )
        )

    logger.info(
        "gemini_scan_completed",
        job_id=str(job.id),
        vendor=extraction.vendor_name,
        items=len(extraction.items),
        confidence=extraction.confidence_score,
    )


# ── コンテンツ構築ヘルパー ─────────────────────────────────────────────────────

def _build_contents(file_type: ScanJobFileType, file_path: Path) -> list[Any]:
    if file_type == ScanJobFileType.pdf:
        return _pdf_to_contents(file_path)
    elif file_type == ScanJobFileType.image:
        return _image_to_contents(file_path)
    elif file_type == ScanJobFileType.excel:
        return _excel_to_contents(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")


def _extraction_prompt() -> str:
    return (
        "あなたは建設業の業者見積書・請求書を読み取るアシスタントです。\n"
        "この書類から以下の情報を抽出し、JSON形式で返してください。\n\n"
        "フィールド定義:\n"
        "- vendor_name: 発行業者（会社）名。不明は null\n"
        "- quoted_date: 見積日・請求日（YYYY-MM-DD形式）。不明は null\n"
        "- items: 明細行リスト\n"
        "  - row_no: 1始まりの行番号\n"
        "  - item_name: 品目・工事名\n"
        "  - spec: 仕様・規格（なければ null）\n"
        "  - unit: 単位（個・m・式 など）\n"
        "  - quantity: 数量（数値）\n"
        "  - unit_price: 単価（円・数値）\n"
        "  - amount: 金額（円・数値）\n"
        "  - confidence: この行の読み取り信頼度（0.0〜1.0）\n"
        "- subtotal: 税抜小計（数値）\n"
        "- tax: 消費税額（数値）\n"
        "- total: 税込合計（数値）\n"
        "- confidence_score: 全体の信頼度（0.0〜1.0）\n\n"
        "注意: 数値フィールドはカンマや円記号を除いた数値型で返すこと。"
        "読み取れない場合は null にしてください。"
    )


def _pdf_to_contents(file_path: Path) -> list[Any]:
    from google.genai import types
    from pdf2image import convert_from_path

    images = convert_from_path(str(file_path), dpi=200, fmt="jpeg")
    contents: list[Any] = []

    for i, img in enumerate(images[:_PDF_MAX_PAGES]):
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        contents.append(
            types.Part.from_bytes(data=buf.getvalue(), mime_type="image/jpeg")
        )
        logger.debug("pdf_page_converted", page=i + 1, total=min(len(images), _PDF_MAX_PAGES))

    contents.append(_extraction_prompt())
    return contents


def _image_to_contents(file_path: Path) -> list[Any]:
    from google.genai import types

    suffix = file_path.suffix.lower()
    mime_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png"}
    mime_type = mime_map.get(suffix, "image/jpeg")

    with open(file_path, "rb") as f:
        data = f.read()

    return [types.Part.from_bytes(data=data, mime_type=mime_type), _extraction_prompt()]


def _excel_to_contents(file_path: Path) -> list[Any]:
    import openpyxl

    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    lines: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== シート: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append("\t".join(cells))

    text = "\n".join(lines)
    return [f"以下はExcel見積書から抽出したテキストです:\n\n{text}\n\n{_extraction_prompt()}"]
