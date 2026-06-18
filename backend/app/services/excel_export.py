"""Excel帳票生成サービス。openpyxl でデータを埋め込む。"""
from __future__ import annotations

import io
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── ロゴパス ──────────────────────────────────────────────────────────────────
_HERE = Path(__file__).parent
LOGO_PATH = _HERE.parent / "templates" / "images" / "clap_logo.png"

# ── 会社情報（CLAUDE.md より） ───────────────────────────────────────────────
COMPANY_NAME = "株式会社クラップ"
COMPANY_NAME_EN = "CLAP CORPORATION"
COMPANY_ADDRESS = "〒913-0043 福井県坂井市三国町錦3-4-2"
COMPANY_TEL = "TEL: 0776-81-8330"
COMPANY_FAX = "FAX: 0776-81-8331"
COMPANY_CEO = "代表取締役　奴間 正人"
COMPANY_REG_NO = "登録番号 T5210001007332"


# ── スタイルヘルパー ──────────────────────────────────────────────────────────

def _thin() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_box() -> Border:
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor="1F4E79")


def _sub_fill() -> PatternFill:
    return PatternFill("solid", fgColor="BDD7EE")


def _apply_table_header(ws: Any, row: int, cols: list[tuple[str, int]]) -> int:
    """テーブルヘッダを書いて次の行番号を返す。"""
    col = 1
    for label, width in cols:
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()
        ws.column_dimensions[get_column_letter(col)].width = width
        col += 1
    ws.row_dimensions[row].height = 20
    return row + 1


def _fmt(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float | int):
        return f"{int(v):,}"
    if isinstance(v, date):
        return v.strftime("%Y/%m/%d")
    return str(v)


def _money_cell(ws: Any, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = int(value)
        cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = _thin()


def _company_block(ws: Any, start_row: int, start_col: int) -> None:
    """右側に会社情報ブロックを書き込む。"""
    info = [
        COMPANY_NAME,
        COMPANY_NAME_EN,
        COMPANY_ADDRESS,
        COMPANY_TEL + "  " + COMPANY_FAX,
        COMPANY_CEO,
        COMPANY_REG_NO,
    ]
    for i, text in enumerate(info):
        cell = ws.cell(row=start_row + i, column=start_col, value=text)
        cell.font = Font(size=9 if i > 0 else 11, bold=(i == 0))
        cell.alignment = Alignment(horizontal="right", vertical="center")


def _insert_logo(ws: Any, anchor: str, width_px: int = 110, height_px: int = 55) -> None:
    """ロゴ画像をシートに貼り付ける（ファイルが存在しない場合はスキップ）。"""
    if not LOGO_PATH.exists():
        return
    try:
        img = XLImage(str(LOGO_PATH))
        img.width = width_px
        img.height = height_px
        ws.add_image(img, anchor)
    except Exception:
        pass  # 画像挿入失敗は無視して帳票生成を継続


# ── 見積書 ────────────────────────────────────────────────────────────────────

_QUOTE_COL_WIDTHS = [4, 28, 18, 6, 10, 14, 14, 16]
_QUOTE_TABLE_COLS = [
    ("No", 4), ("工事項目", 28), ("仕様", 18),
    ("単位", 6), ("数量", 10), ("単価", 14), ("金額", 14), ("備考", 16),
]


def _write_quote_header(ws: Any, quote: Any, project: Any) -> int:
    """見積書ヘッダ（ロゴ・タイトル・会社情報・工事名・合計額）を書き、次の行番号を返す。"""
    for i, w in enumerate(_QUOTE_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _insert_logo(ws, anchor="A1", width_px=110, height_px=55)

    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=4, value="見　積　書")
    title.font = Font(bold=True, size=20)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("D1:H1")

    _company_block(ws, 2, 5)

    ws.cell(row=2, column=1, value="見積番号").font = Font(size=9, bold=True)
    ws.cell(row=2, column=2, value=quote.quote_number or "")
    ws.cell(row=2, column=3, value="発行日").font = Font(size=9, bold=True)
    ws.cell(row=2, column=4, value=_fmt(quote.issue_date))

    ws.cell(row=3, column=1, value="有効期限").font = Font(size=9, bold=True)
    ws.cell(row=3, column=2, value=f"発行日より {quote.validity_days or 30} 日間")

    ws.row_dimensions[4].height = 4
    ws.cell(row=5, column=1, value=f"{project.client_name or ''} 御中")
    ws.cell(row=5, column=1).font = Font(bold=True, size=13)
    ws.merge_cells("A5:D5")

    ws.cell(row=6, column=1, value="下記の通り御見積申し上げます。")
    ws.cell(row=6, column=1).font = Font(size=9, italic=True)
    ws.merge_cells("A6:D6")

    ws.row_dimensions[7].height = 4
    ws.cell(row=8, column=1, value="工　事　名").font = Font(size=9, bold=True)
    ws.cell(row=8, column=2, value=project.project_name or "")
    ws.merge_cells("B8:H8")

    ws.cell(row=9, column=1, value="工事場所").font = Font(size=9, bold=True)
    ws.cell(row=9, column=2, value=quote.project_location_snapshot or project.project_location or "")
    ws.merge_cells("B9:H9")

    ws.cell(row=10, column=1, value="工　　期").font = Font(size=9, bold=True)
    period_text = ""
    if quote.period_start and quote.period_end:
        period_text = f"{_fmt(quote.period_start)} ～ {_fmt(quote.period_end)}"
    ws.cell(row=10, column=2, value=period_text)
    ws.merge_cells("B10:H10")

    ws.cell(row=11, column=1, value="支払条件").font = Font(size=9, bold=True)
    ws.cell(row=11, column=2, value=quote.payment_condition or "")
    ws.merge_cells("B11:H11")

    ws.row_dimensions[12].height = 4
    ws.row_dimensions[13].height = 28
    total_label = ws.cell(row=13, column=1, value="御　見　積　金　額")
    total_label.font = Font(bold=True, size=11)
    total_label.fill = _sub_fill()
    total_label.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A13:D13")

    total_cell = ws.cell(row=13, column=5, value=int(quote.total_amount or 0))
    total_cell.font = Font(bold=True, size=14)
    total_cell.number_format = "¥#,##0"
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    for cell in [ws.cell(row=13, column=c) for c in range(1, 9)]:
        cell.border = _thick_box()

    tax_note = ws.cell(row=13, column=6, value="（消費税10%込）")
    tax_note.font = Font(size=9, color="666666")
    tax_note.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("F13:H13")

    ws.row_dimensions[14].height = 4
    return 15  # 次の行


def _write_item_rows(ws: Any, start_row: int, items: list[Any]) -> int:
    """明細行を書き、次の行番号を返す。"""
    data_row = _apply_table_header(ws, start_row, _QUOTE_TABLE_COLS)
    filled = [i for i in items if i.item_name or i.amount]
    for idx, item in enumerate(filled, 1):
        r = data_row
        ws.row_dimensions[r].height = 16
        ws.cell(r, 1, idx).border = _thin()
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")

        c2 = ws.cell(r, 2, item.item_name or "")
        c2.border = _thin()
        c2.alignment = Alignment(vertical="center", wrap_text=True)

        c3 = ws.cell(r, 3, item.spec or "")
        c3.border = _thin()
        c3.alignment = Alignment(vertical="center", wrap_text=True)

        c4 = ws.cell(r, 4, item.unit or "")
        c4.border = _thin()
        c4.alignment = Alignment(horizontal="center", vertical="center")

        c5 = ws.cell(r, 5, float(item.quantity) if item.quantity else None)
        c5.border = _thin()
        c5.number_format = "#,##0.0##"
        c5.alignment = Alignment(horizontal="right", vertical="center")

        _money_cell(ws, r, 6, item.unit_price)
        _money_cell(ws, r, 7, item.amount or (
            (float(item.quantity or 0) * float(item.unit_price or 0))
            if item.quantity and item.unit_price else None
        ))

        c8 = ws.cell(r, 8, item.remarks or "")
        c8.border = _thin()
        c8.alignment = Alignment(vertical="center")

        data_row += 1
    return data_row


def _write_quote_summary(ws: Any, start_row: int, quote: Any) -> int:
    """小計・消費税・合計行を書き、次の行番号を返す。"""
    data_row = start_row + 1
    for label, val in [
        ("小　計（税抜）", quote.subtotal),
        ("消費税（10%）", quote.tax_amount),
        ("合　計（税込）", quote.total_amount),
    ]:
        ws.row_dimensions[data_row].height = 18
        lc = ws.cell(data_row, 6, label)
        lc.font = Font(bold=True, size=9)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border = _thin()
        ws.merge_cells(f"A{data_row}:F{data_row}")
        _money_cell(ws, data_row, 7, val)
        ws.cell(data_row, 7).font = Font(bold=True)
        ws.cell(data_row, 8).border = _thin()
        data_row += 1
    return data_row


def export_quote_excel(
    quote: Any,
    project: Any,
    items: list[Any],
    sections: list[Any] | None = None,
) -> bytes:
    """見積書をExcel(xlsx)として生成し、bytesで返す。

    大項目（sections）がある場合は 表紙 / 大項目集計 / 大項目別明細 の多シート構成。
    sections が空なら従来の1シート構成にフォールバック。
    """
    wb = Workbook()

    sections_sorted = sorted(sections or [], key=lambda s: s.row_no)

    if sections_sorted:
        # ── 多シート構成 ──────────────────────────────────────────────
        # Sheet1: 表紙（大項目の小計一覧付き）
        ws1 = wb.active
        ws1.title = "表紙"
        next_row = _write_quote_header(ws1, quote, project)

        # 大項目合計一覧を表紙に追加
        ws1.cell(next_row, 1, "大項目別集計").font = Font(bold=True, size=10)
        ws1.merge_cells(f"A{next_row}:H{next_row}")
        next_row += 1
        sec_hdr_row = next_row
        for col, (label, width) in enumerate([("記号", 6), ("大項目名", 34), ("小計", 14)], 1):
            c = ws1.cell(sec_hdr_row, col, label)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = _header_fill()
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin()
            ws1.column_dimensions[get_column_letter(col)].width = width
        next_row += 1
        for sec in sections_sorted:
            ws1.cell(next_row, 1, sec.section_letter).border = _thin()
            ws1.cell(next_row, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws1.cell(next_row, 1).font = Font(bold=True)
            ws1.cell(next_row, 2, sec.section_name).border = _thin()
            _money_cell(ws1, next_row, 3, sec.amount)
            ws1.merge_cells(f"D{next_row}:H{next_row}")
            next_row += 1

        # 合計行
        _write_quote_summary(ws1, next_row - 1, quote)

        ws1.print_area = f"A1:H{next_row + 8}"
        ws1.page_setup.orientation = "portrait"
        ws1.page_setup.fitToPage = True

        # Sheet2: 大項目集計
        ws2 = wb.create_sheet(title="大項目集計")
        for i, w in enumerate(_QUOTE_COL_WIDTHS, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.cell(1, 1, "大項目集計").font = Font(bold=True, size=14)
        ws2.merge_cells("A1:H1")
        ws2.cell(2, 1, f"見積番号：{quote.quote_number or ''}").font = Font(size=9)
        ws2.cell(2, 4, f"案件名：{project.project_name}").font = Font(size=9)
        r = 4
        r = _apply_table_header(ws2, r, [("記号", 6), ("大項目名", 40), ("小計（税抜）", 18), ("割合", 10)])
        subtotal_total = float(quote.subtotal or 0)
        for sec in sections_sorted:
            ws2.cell(r, 1, sec.section_letter).border = _thin()
            ws2.cell(r, 1).alignment = Alignment(horizontal="center")
            ws2.cell(r, 1).font = Font(bold=True)
            ws2.cell(r, 2, sec.section_name).border = _thin()
            _money_cell(ws2, r, 3, sec.amount)
            pct = (float(sec.amount or 0) / subtotal_total * 100) if subtotal_total else 0
            c4 = ws2.cell(r, 4, round(pct, 1))
            c4.number_format = "0.0%"
            c4.border = _thin()
            c4.alignment = Alignment(horizontal="right")
            r += 1
        r += 1
        for label, val in [
            ("小計（税抜）", quote.subtotal),
            ("消費税（10%）", quote.tax_amount),
            ("合計（税込）", quote.total_amount),
        ]:
            lc = ws2.cell(r, 2, label)
            lc.font = Font(bold=True, size=9)
            lc.border = _thin()
            _money_cell(ws2, r, 3, val)
            ws2.cell(r, 3).font = Font(bold=True)
            ws2.cell(r, 1).border = _thin()
            ws2.cell(r, 4).border = _thin()
            r += 1

        ws2.print_area = f"A1:D{r}"
        ws2.page_setup.fitToPage = True

        # Sheet3+: 大項目別明細
        item_by_section: dict[str, list[Any]] = {}
        unsectioned: list[Any] = []
        for item in items:
            sid = str(item.section_id) if item.section_id else None
            if sid:
                item_by_section.setdefault(sid, []).append(item)
            else:
                unsectioned.append(item)

        for sec in sections_sorted:
            sheet_name = f"{sec.section_letter}.{sec.section_name}"[:31]
            ws_sec = wb.create_sheet(title=sheet_name)
            for i, w in enumerate(_QUOTE_COL_WIDTHS, 1):
                ws_sec.column_dimensions[get_column_letter(i)].width = w

            ws_sec.cell(1, 1, f"{sec.section_letter}. {sec.section_name}").font = Font(bold=True, size=13)
            ws_sec.merge_cells("A1:H1")
            ws_sec.cell(2, 1, f"見積番号：{quote.quote_number or ''}  案件：{project.project_name}").font = Font(size=9, color="666666")
            ws_sec.merge_cells("A2:H2")
            ws_sec.row_dimensions[3].height = 4

            sec_items = sorted(item_by_section.get(str(sec.id), []), key=lambda x: x.row_no)
            dr = _write_item_rows(ws_sec, 4, sec_items)

            # 小計行
            dr += 1
            sec_subtotal = sum(float(i.amount or 0) for i in sec_items)
            lc = ws_sec.cell(dr, 6, "小　計")
            lc.font = Font(bold=True, size=9)
            lc.alignment = Alignment(horizontal="right")
            lc.border = _thin()
            ws_sec.merge_cells(f"A{dr}:F{dr}")
            _money_cell(ws_sec, dr, 7, sec_subtotal)
            ws_sec.cell(dr, 7).font = Font(bold=True)
            ws_sec.cell(dr, 8).border = _thin()

            ws_sec.print_area = f"A1:H{dr + 2}"
            ws_sec.page_setup.fitToPage = True

        if unsectioned:
            ws_other = wb.create_sheet(title="大項目未分類")
            for i, w in enumerate(_QUOTE_COL_WIDTHS, 1):
                ws_other.column_dimensions[get_column_letter(i)].width = w
            ws_other.cell(1, 1, "大項目未分類").font = Font(bold=True, size=13)
            ws_other.merge_cells("A1:H1")
            ws_other.row_dimensions[2].height = 4
            dr = _write_item_rows(ws_other, 3, unsectioned)
            ws_other.print_area = f"A1:H{dr + 2}"
            ws_other.page_setup.fitToPage = True

    else:
        # ── 従来の1シート構成（大項目なし） ──────────────────────────
        ws = wb.active
        ws.title = "見積書"
        next_row = _write_quote_header(ws, quote, project)
        data_row = _write_item_rows(ws, next_row, items)
        data_row = _write_quote_summary(ws, data_row, quote)

        data_row += 1
        if quote.conditions_text:
            ws.cell(data_row, 1, "工事条件等").font = Font(bold=True, size=9)
            data_row += 1
            for line in quote.conditions_text.split("\n"):
                c = ws.cell(data_row, 1, line)
                c.font = Font(size=9)
                ws.merge_cells(f"A{data_row}:H{data_row}")
                data_row += 1

        if quote.remarks:
            data_row += 1
            ws.cell(data_row, 1, "備考").font = Font(bold=True, size=9)
            data_row += 1
            ws.cell(data_row, 1, quote.remarks).font = Font(size=9)
            ws.merge_cells(f"A{data_row}:H{data_row}")

        ws.print_area = f"A1:H{data_row + 2}"
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


