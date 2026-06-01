"""既存Excel工事台帳の全シート対応インポートサービス。

シートマッピング:
  工事台帳    → Project（案件）
  QCDS        → QCDS + QCDSDirectWork[]（直接工事費）
  表紙        → Quote ヘッダ（顧客見積書表紙）
  内訳書      → QuoteSection[] + QuoteItem[]（顧客見積明細）
  注文書・請書 → Order + Acknowledgment
  請求書      → Invoice + InvoiceItem[]

「記入例」シートおよび非案件シートは自動除外される。
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.enums import ContractType, OrderType


# ---------------------------------------------------------------------------
# 関連シート解析データクラス
# ---------------------------------------------------------------------------

@dataclass
class ExcelDirectWork:
    """QCDSシートの直接工事費1行。"""
    row_no: int
    vendor_name: str | None
    budget_amount: float | None
    category: str = "subcontract"  # 外注(subcontract) / 資材(material) / その他(other)


@dataclass
class ExcelQuoteItem:
    """内訳書の明細1行。"""
    row_no: int
    item_name: str | None = None
    spec: str | None = None
    unit: str | None = None
    quantity: float | None = None
    unit_price: float | None = None
    amount: float | None = None
    remarks: str | None = None


@dataclass
class ExcelQuoteSection:
    """内訳書の大項目。"""
    letter: str
    name: str
    row_no: int
    items: list[ExcelQuoteItem] = field(default_factory=list)


@dataclass
class ExcelQuoteData:
    """表紙・内訳書から収集した見積書データ。"""
    quote_number: str | None = None
    issue_date: date | None = None
    project_name_snapshot: str | None = None
    project_location_snapshot: str | None = None
    validity_days: int = 30
    payment_condition: str | None = None
    period_text: str | None = None
    remarks: str | None = None
    subtotal: float | None = None
    tax_amount: float | None = None
    total_amount: float | None = None
    discount_amount: float | None = None
    sections: list[ExcelQuoteSection] = field(default_factory=list)
    unsectioned_items: list[ExcelQuoteItem] = field(default_factory=list)


@dataclass
class ExcelOrderData:
    """注文書・請書シートから収集した注文書データ。"""
    amount_excl_tax: float | None = None
    tax_amount: float | None = None
    total_amount: float | None = None
    payment_condition: str | None = None
    period_start_text: str | None = None
    period_end_text: str | None = None
    client_company: str | None = None


@dataclass
class ExcelInvoiceItem:
    """請求書シートの明細1行。"""
    row_no: int
    item_name: str | None = None
    amount: float | None = None


@dataclass
class ExcelInvoiceData:
    """請求書シートから収集した請求書データ。"""
    current_purchase: float | None = None
    tax_amount: float | None = None
    total_amount: float | None = None
    items: list[ExcelInvoiceItem] = field(default_factory=list)


# ---------------------------------------------------------------------------
# 案件行クラス（1シート = 1案件）
# ---------------------------------------------------------------------------

class ExcelImportRow:
    """1シート = 1案件として抽出したデータ（関連シートデータ含む）。"""

    def __init__(self) -> None:
        # --- 案件基本情報（工事台帳シート）---
        self.project_number: str | None = None
        self.project_name: str | None = None
        self.project_location: str | None = None
        self.client_name: str | None = None
        self.original_client_name: str | None = None
        self.project_price: float | None = None
        self.payment_condition: str | None = None
        self.client_contact_company: str | None = None
        self.client_contact_person: str | None = None
        self.order_type: OrderType | None = None
        self.contract_type: ContractType | None = None
        self.period_quote_start: date | None = None
        self.period_quote_end: date | None = None
        self.period_contract_start: date | None = None
        self.period_contract_end: date | None = None
        self.period_actual_start: date | None = None
        self.period_actual_end: date | None = None
        # --- 関連シートデータ ---
        self.qcds_direct_works: list[ExcelDirectWork] = []
        # B/C セクション経費金額の上書き値（key=system_key, value=Excel実績金額）
        # None でなく 0.0 が入った場合は「Excelで明示的に0」として amount_override に使用する
        self.qcds_expense_overrides: dict[str, float] = {}
        self.quote: ExcelQuoteData | None = None
        self.order: ExcelOrderData | None = None
        self.invoice: ExcelInvoiceData | None = None


# ---------------------------------------------------------------------------
# セル値ヘルパー
# ---------------------------------------------------------------------------

def _cell(ws: Worksheet, coord: str) -> Any:
    """セル値を返す。None・空文字・数式エラーは None に正規化。"""
    v = ws[coord].value
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?"):
        return None
    return v


def _str(v: Any) -> str | None:
    """セル値を文字列に変換。空・None は None。"""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_int(v: Any) -> int | None:
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _build_date(year: Any, month: Any, day: Any) -> date | None:
    try:
        y = int(year) if year else None
        m = int(month) if month else None
        d = int(day) if day else None
        if y and m and d and 1 <= m <= 12 and 1 <= d <= 31:
            return date(y, m, d)
    except (TypeError, ValueError):
        pass
    return None


def _parse_date_row(ws: Worksheet, row: int) -> tuple[date | None, date | None]:
    """工期行(row)の開始・終了日を返す。
    書式: L(年) N(月) P(日)  〜  X(年) Z(月) AB(日)
    """
    raw_start = _cell(ws, f"L{row}")
    raw_end = _cell(ws, f"X{row}")

    def parse_slash(raw: Any, month_col: str, day_col: str) -> date | None:
        if raw is None:
            return None
        s = str(raw).strip()
        m = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})", s)
        if m:
            return _build_date(m.group(1), m.group(2), m.group(3))
        m2 = re.match(r"(\d{4})/", s)
        if m2:
            year = m2.group(1)
            mo = _cell(ws, f"{month_col}{row}")
            da = _cell(ws, f"{day_col}{row}")
            return _build_date(year, mo, da)
        try:
            import datetime as dt
            serial = int(raw)
            if serial > 0:
                base = dt.date(1899, 12, 30)
                return base + dt.timedelta(days=serial)
        except (TypeError, ValueError):
            pass
        return None

    start = parse_slash(raw_start, "N", "P")
    end = parse_slash(raw_end, "Z", "AB")
    return start, end


def _parse_order_type(ws: Worksheet) -> tuple[OrderType | None, ContractType | None]:
    """L22 の '民間・官庁' と T22 の '元請・下請' を解析。"""
    val22 = _cell(ws, "L22")
    val_t22 = _cell(ws, "T22")

    order_type: OrderType | None = None
    contract_type: ContractType | None = None

    if val22:
        s = str(val22)
        if "民間" in s:
            order_type = OrderType.private
        elif "官庁" in s:
            order_type = OrderType.government

    if val_t22:
        s = str(val_t22)
        if "元請" in s:
            contract_type = ContractType.prime
        elif "下請" in s:
            contract_type = ContractType.sub

    return order_type, contract_type


# ---------------------------------------------------------------------------
# 工事台帳シート解析
# ---------------------------------------------------------------------------

def parse_worksheet(ws: Worksheet) -> ExcelImportRow | None:
    """1ワークシートから案件データを抽出する。データが無ければ None。"""
    project_number = _cell(ws, "L11")
    project_name = _cell(ws, "L12")

    if not project_name:
        return None

    row = ExcelImportRow()
    row.project_number = str(project_number).strip() if project_number else None
    row.project_name = str(project_name).strip()
    row.project_location = str(_cell(ws, "L14") or "").strip() or None
    row.client_name = str(_cell(ws, "L16") or "").strip() or None
    row.original_client_name = str(_cell(ws, "C18") or "").strip() or None
    if row.original_client_name and row.original_client_name in ("(元発注者）", "元発注者"):
        row.original_client_name = None
    row.project_price = _to_float(_cell(ws, "BB11"))
    row.payment_condition = str(_cell(ws, "L23") or "").strip() or None
    row.client_contact_company = str(_cell(ws, "Q34") or "").strip() or None
    row.client_contact_person = str(_cell(ws, "Q35") or "").strip() or None

    row.order_type, row.contract_type = _parse_order_type(ws)
    row.period_quote_start, row.period_quote_end = _parse_date_row(ws, 19)
    row.period_contract_start, row.period_contract_end = _parse_date_row(ws, 20)
    row.period_actual_start, row.period_actual_end = _parse_date_row(ws, 21)

    return row


# ---------------------------------------------------------------------------
# QCDSシート解析
# ---------------------------------------------------------------------------

def _parse_qcds_sheet(ws: Worksheet) -> tuple[list[ExcelDirectWork], dict[str, float]]:
    """QCDSシートの直接工事費（A）と経費項目（B）の金額を読む。

    A セクション: rows 15-43, B=NO, C=業者名, K=予算
    B セクション: rows 50-60, AN列=金額（Excel計算結果）
    """
    # ── A セクション: 直接工事費 ───────────────────────────────────────────
    works: list[ExcelDirectWork] = []
    for rn in range(15, 44):
        b_val = _cell(ws, f"B{rn}")
        c_val = _cell(ws, f"C{rn}")
        k_val = _cell(ws, f"K{rn}")

        vendor = _str(c_val)
        if not vendor:
            continue

        row_no = _to_int(b_val) or (rn - 14)
        budget = _to_float(k_val)
        works.append(ExcelDirectWork(row_no=row_no, vendor_name=vendor, budget_amount=budget))

    # ── B セクション: 経費項目 (金額列 = AN = col 40) ─────────────────────
    # 行番号 → system_key マッピング（QCDSExpenseItem.system_key と対応）
    _ROW_TO_KEY: dict[int, str] = {
        50: "labor_insurance",
        51: "construction_insurance",
        52: "stamp_cost",
        53: "receipt_cost",
        54: "special_insurance",
        58: "site_personnel_cost",
        59: "construction_dept_overhead",
        60: "shared_overhead",
    }
    expense_overrides: dict[str, float] = {}
    for rn, key in _ROW_TO_KEY.items():
        v = _to_float(_cell(ws, f"AN{rn}"))
        if v is not None:
            expense_overrides[key] = v

    # 事務用品(row55) + 雑費(row56) → fixed_overhead
    office = _to_float(_cell(ws, "AN55"))
    misc = _to_float(_cell(ws, "AN56"))
    if office is not None or misc is not None:
        expense_overrides["fixed_overhead"] = (office or 0.0) + (misc or 0.0)

    return works, expense_overrides


# ---------------------------------------------------------------------------
# 表紙シート解析
# ---------------------------------------------------------------------------

def _parse_hyoshi_sheet(ws: Worksheet) -> ExcelQuoteData:
    """表紙シートから見積書ヘッダを読む。"""
    data = ExcelQuoteData()

    # 工事番号 (R1)
    data.quote_number = _str(_cell(ws, "R1"))

    # 発行日: O2(年) R2(月) U2(日)
    year = _to_int(_cell(ws, "O2"))
    month = _to_int(_cell(ws, "R2"))
    day = _to_int(_cell(ws, "U2"))
    data.issue_date = _build_date(year, month, day)

    # 税込合計 (C14)
    data.total_amount = _to_float(_cell(ws, "C14"))

    # 工事名・場所
    data.project_name_snapshot = _str(_cell(ws, "C18"))
    data.project_location_snapshot = _str(_cell(ws, "C20"))

    # 支払条件 (C24)
    data.payment_condition = _str(_cell(ws, "C24"))

    # 有効期限 C22: '30日間' → 30
    validity_raw = _str(_cell(ws, "C22"))
    if validity_raw:
        m = re.search(r"(\d+)", validity_raw)
        if m:
            data.validity_days = int(m.group(1))

    # 工期 (C26 ～ F26) — テキスト保存
    p_start = _str(_cell(ws, "C26"))
    p_end = _str(_cell(ws, "F26"))
    parts = [p for p in [p_start, p_end] if p]
    if parts:
        data.period_text = " ～ ".join(parts)

    # 備考 (D28-D31)
    remarks_parts = [_str(_cell(ws, f"D{r}")) for r in range(28, 32)]
    remarks_parts = [p for p in remarks_parts if p]
    if remarks_parts:
        data.remarks = "\n".join(remarks_parts)

    return data


# ---------------------------------------------------------------------------
# 内訳書シート解析
# ---------------------------------------------------------------------------

_SKIP_NAMES = frozenset({
    "計", "消費税", "合　計", "合計",
    "小　計", "小計", "出精値引き",
})

_FULLWIDTH_MAP = str.maketrans(
    "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ",
    "ABCDEFGHIJKLMNOPQRSTUVWXYZ",
)


def _is_section_letter(v: Any) -> bool:
    """単一英字（半角・全角）ならセクション記号と判定。"""
    if v is None:
        return False
    s = str(v).strip().translate(_FULLWIDTH_MAP).upper()
    return len(s) == 1 and "A" <= s <= "Z"


def _normalize_letter(v: Any) -> str:
    return str(v).strip().translate(_FULLWIDTH_MAP).upper()


def _parse_naiwake_sheet(
    ws: Worksheet,
) -> tuple[
    list[ExcelQuoteSection],
    list[ExcelQuoteItem],
    float | None,  # subtotal
    float | None,  # tax
    float | None,  # total
    float | None,  # discount
]:
    """内訳書シートから大項目・明細を読む。

    サマリー行（rows 2-20）から大項目を作成し、
    詳細行（rows 21+）から明細を作成する。
    詳細行が無い場合はサマリー行から代替アイテムを生成する。
    """
    sections_map: dict[str, ExcelQuoteSection] = {}
    summary_items: dict[str, ExcelQuoteItem] = {}
    unsectioned: list[ExcelQuoteItem] = []
    subtotal = tax = total = discount = None

    # Phase 1: サマリーブロック (rows 2-20)
    sec_row_no = 1
    for rn in range(2, 21):
        a = _cell(ws, f"A{rn}")
        b = _cell(ws, f"B{rn}")
        g_val = _cell(ws, f"G{rn}")
        b_str = _str(b) or ""

        if _is_section_letter(a) and b_str and b_str not in _SKIP_NAMES:
            letter = _normalize_letter(a)
            sections_map[letter] = ExcelQuoteSection(
                letter=letter,
                name=b_str,
                row_no=sec_row_no,
            )
            sec_row_no += 1
            # サマリー行をアイテムの代替として保持
            summary_items[letter] = ExcelQuoteItem(
                row_no=1,
                item_name=b_str,
                spec=_str(_cell(ws, f"C{rn}")),
                unit=_str(_cell(ws, f"D{rn}")) or "式",
                quantity=_to_float(_cell(ws, f"E{rn}")) or 1,
                unit_price=_to_float(_cell(ws, f"F{rn}")),
                amount=_to_float(g_val),
                remarks=_str(_cell(ws, f"H{rn}")),
            )
            continue

        if b_str == "出精値引き":
            v = _to_float(g_val)
            if v is not None:
                discount = abs(v)
        elif b_str == "計":
            subtotal = _to_float(g_val)
        elif "消費税" in b_str:
            tax = _to_float(g_val)
        elif "合" in b_str and "計" in b_str:
            total = _to_float(g_val)

    # Phase 2: 詳細ブロック (rows 21+)
    current_letter: str | None = None
    detail_row_no = 1
    max_row = min((ws.max_row or 200) + 1, 500)
    for rn in range(21, max_row):
        a = _cell(ws, f"A{rn}")
        b = _cell(ws, f"B{rn}")
        b_str = _str(b) or ""

        if _is_section_letter(a):
            current_letter = _normalize_letter(a)
            continue  # セクション見出し行はアイテムにしない

        if not b_str or b_str in _SKIP_NAMES:
            continue

        # セクション名と一致する行はスキップ（見出し行の重複）
        if current_letter and current_letter in sections_map:
            if b_str == sections_map[current_letter].name:
                continue

        item = ExcelQuoteItem(
            row_no=detail_row_no,
            item_name=b_str,
            spec=_str(_cell(ws, f"C{rn}")),
            unit=_str(_cell(ws, f"D{rn}")),
            quantity=_to_float(_cell(ws, f"E{rn}")),
            unit_price=_to_float(_cell(ws, f"F{rn}")),
            amount=_to_float(_cell(ws, f"G{rn}")),
            remarks=_str(_cell(ws, f"H{rn}")),
        )
        detail_row_no += 1

        if current_letter and current_letter in sections_map:
            sections_map[current_letter].items.append(item)
        else:
            unsectioned.append(item)

    # 詳細行が無い場合はサマリー行をアイテムとして使用
    has_detail = any(s.items for s in sections_map.values()) or bool(unsectioned)
    if not has_detail:
        for letter, sec in sections_map.items():
            if letter in summary_items:
                sec.items.append(summary_items[letter])

    return list(sections_map.values()), unsectioned, subtotal, tax, total, discount


# ---------------------------------------------------------------------------
# 注文書・請書シート解析
# ---------------------------------------------------------------------------

def _parse_order_sheet(ws: Worksheet) -> ExcelOrderData:
    """注文書・請書シートから注文書データを読む。"""
    data = ExcelOrderData()

    # 工事代金 (D24=税抜, D27=税, D30=合計)
    data.amount_excl_tax = _to_float(_cell(ws, "D24"))
    data.tax_amount = _to_float(_cell(ws, "D27"))
    data.total_amount = _to_float(_cell(ws, "D30"))
    data.payment_condition = _str(_cell(ws, "C38"))

    # 工期 (C34 ～ F34)
    data.period_start_text = _str(_cell(ws, "C34"))
    data.period_end_text = _str(_cell(ws, "F34"))

    # 発注者（注文請書の宛先 A55、なければ注文書の宛先 A7 から）
    data.client_company = _str(_cell(ws, "A55")) or _str(_cell(ws, "A7"))

    return data


# ---------------------------------------------------------------------------
# 請求書シート解析
# ---------------------------------------------------------------------------

_INVOICE_SKIP_NAMES = frozenset({
    "計", "消費税（10％）", "合　　計", "合計",
    "日付", "工事名・備考", "金　　額",
})


def _parse_invoice_sheet(ws: Worksheet) -> ExcelInvoiceData:
    """請求書シートから請求書データを読む。"""
    data = ExcelInvoiceData()

    # 当月明細ヘッダ行 (L19=税抜, P19=税, T19=合計)
    data.current_purchase = _to_float(_cell(ws, "L19"))
    data.tax_amount = _to_float(_cell(ws, "P19"))
    data.total_amount = _to_float(_cell(ws, "T19"))

    # 明細行 (rows 24-38, E=工事名, O=金額)
    item_row_no = 1
    for rn in range(24, 40):
        e_val = _cell(ws, f"E{rn}")
        o_val = _cell(ws, f"O{rn}")
        e_str = _str(e_val) or ""
        if not e_str or e_str in _INVOICE_SKIP_NAMES:
            continue
        amount = _to_float(o_val)
        if e_str or amount:
            data.items.append(ExcelInvoiceItem(
                row_no=item_row_no,
                item_name=e_str or None,
                amount=amount,
            ))
            item_row_no += 1

    return data


# ---------------------------------------------------------------------------
# メイン解析関数
# ---------------------------------------------------------------------------

_NON_PROJECT_SHEETS = frozenset({
    "qcds", "見積条件書", "表紙", "内訳書",
    "注文書・請書", "注文書", "請書", "請求書", "印紙税額算定表", "印紙税額算定表1",
})

# 部分一致で除外するキーワード（上記完全一致に加えて）
_NON_PROJECT_KEYWORDS = ("注文書", "請書", "請求書", "qcds", "見積条件", "印紙税")


def parse_excel(file_bytes: bytes) -> list[ExcelImportRow]:
    """Excelバイト列を解析して案件データリストを返す。

    各シートを以下のように処理する:
    - 「記入例」シート: 除外
    - 非案件シート(QCDS/表紙等): 関連データとして別途解析
    - 工事台帳シート: 1件の案件として解析
    """
    import io

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 名前付きシートのルックアップ（関連シート）
    named_sheets: dict[str, Worksheet] = {ws.title.strip(): ws for ws in wb.worksheets}

    # 部分一致でシートを検索するヘルパー（名前に空白・表記揺れがあっても対応）
    def _find_sheet(*candidates: str) -> Worksheet | None:
        for c in candidates:
            if c in named_sheets:
                return named_sheets[c]
        # 部分一致フォールバック
        for name, ws_ in named_sheets.items():
            if any(c in name for c in candidates):
                return ws_
        return None

    qcds_ws = _find_sheet("QCDS")
    hyoshi_ws = _find_sheet("表紙")
    naiwake_ws = _find_sheet("内訳書")
    order_ws = _find_sheet("注文書・請書", "注文書請書")
    invoice_ws = _find_sheet("請求書")

    # 関連シートデータ解析（ブック全体で共通）
    direct_works, expense_overrides = _parse_qcds_sheet(qcds_ws) if qcds_ws else ([], {})

    quote_data: ExcelQuoteData | None = None
    if hyoshi_ws or naiwake_ws:
        quote_data = ExcelQuoteData()
        if hyoshi_ws:
            hq = _parse_hyoshi_sheet(hyoshi_ws)
            quote_data.quote_number = hq.quote_number
            quote_data.issue_date = hq.issue_date
            quote_data.project_name_snapshot = hq.project_name_snapshot
            quote_data.project_location_snapshot = hq.project_location_snapshot
            quote_data.validity_days = hq.validity_days
            quote_data.payment_condition = hq.payment_condition
            quote_data.period_text = hq.period_text
            quote_data.remarks = hq.remarks
            quote_data.total_amount = hq.total_amount
        if naiwake_ws:
            sections, unsec, subtotal, tax, total, discount = _parse_naiwake_sheet(naiwake_ws)
            quote_data.sections = sections
            quote_data.unsectioned_items = unsec
            if subtotal is not None:
                quote_data.subtotal = subtotal
            if tax is not None:
                quote_data.tax_amount = tax
            if total is not None and quote_data.total_amount is None:
                quote_data.total_amount = total
            if discount is not None:
                quote_data.discount_amount = discount

    order_data = _parse_order_sheet(order_ws) if order_ws else None
    invoice_data = _parse_invoice_sheet(invoice_ws) if invoice_ws else None

    # 案件シート解析
    results: list[ExcelImportRow] = []
    for ws in wb.worksheets:
        title = ws.title.strip()
        title_norm = title.lower()
        if "記入例" in title:
            continue
        if title_norm in _NON_PROJECT_SHEETS:
            continue
        # 部分一致除外：注文書・請書・請求書などのシートはプロジェクトシートとして扱わない
        if any(kw in title_norm for kw in _NON_PROJECT_KEYWORDS):
            continue

        row = parse_worksheet(ws)
        if row is None:
            continue

        row.qcds_direct_works = direct_works
        row.qcds_expense_overrides = expense_overrides
        row.quote = quote_data
        row.order = order_data
        row.invoice = invoice_data
        results.append(row)

    return results
