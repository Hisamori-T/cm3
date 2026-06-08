"""Excel帳票生成サービス。openpyxl でデータを埋め込む。"""
from __future__ import annotations

import io
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── ロゴパス ──────────────────────────────────────────────────────────────────
_HERE = Path(__file__).parent
LOGO_PATH = _HERE.parent / "templates" / "images" / "clap_logo.png"

# ── 会社情報（CLAUDE.md より） ───────────────────────────────────────────────
COMPANY_NAME = "株式会社クラップ"
COMPANY_NAME_EN = "CLAP CORPORATION"
COMPANY_ADDRESS = "〒913-0043 福井県坂井市三国町錦3-4-2"
COMPANY_TEL = "TEL: 0776-81-8330"
COMPANY_FAX = "FAX: 0776-81-8331"
COMPANY_CEO = "代表取締役　奴間 正人"
COMPANY_REG_NO = "登録番号 T5210001007332"
COMPANY_BANK = "福井銀行 経田支店 普通 1068586 株式会社クラップ"


# ── スタイルヘルパー ──────────────────────────────────────────────────────────

def _thin() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_box() -> Border:
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor="1F4E79")


def _sub_fill() -> PatternFill:
    return PatternFill("solid", fgColor="BDD7EE")


def _apply_table_header(ws: Any, row: int, cols: list[tuple[str, int]]) -> int:
    """テーブルヘッダを書いて次の行番号を返す。"""
    col = 1
    for label, width in cols:
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin()
        ws.column_dimensions[get_column_letter(col)].width = width
        col += 1
    ws.row_dimensions[row].height = 20
    return row + 1


def _fmt(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float | int):
        return f"{int(v):,}"
    if isinstance(v, date):
        return v.strftime("%Y/%m/%d")
    return str(v)


def _money_cell(ws: Any, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = int(value)
        cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = _thin()


def _company_block(ws: Any, start_row: int, start_col: int) -> None:
    """右側に会社情報ブロックを書き込む。"""
    info = [
        COMPANY_NAME,
        COMPANY_NAME_EN,
        COMPANY_ADDRESS,
        COMPANY_TEL + "  " + COMPANY_FAX,
        COMPANY_CEO,
        COMPANY_REG_NO,
    ]
    for i, text in enumerate(info):
        cell = ws.cell(row=start_row + i, column=start_col, value=text)
        cell.font = Font(size=9 if i > 0 else 11, bold=(i == 0))
        cell.alignment = Alignment(horizontal="right", vertical="center")


def _insert_logo(ws: Any, anchor: str, width_px: int = 110, height_px: int = 55) -> None:
    """ロゴ画像をシートに貼り付ける（ファイルが存在しない場合はスキップ）。"""
    if not LOGO_PATH.exists():
        return
    try:
        img = XLImage(str(LOGO_PATH))
        img.width = width_px
        img.height = height_px
        ws.add_image(img, anchor)
    except Exception:
        pass  # 画像挿入失敗は無視して帳票生成を継続


# ── 見積書 ────────────────────────────────────────────────────────────────────

_QUOTE_COL_WIDTHS = [4, 28, 18, 6, 10, 14, 14, 16]
_QUOTE_TABLE_COLS = [
    ("No", 4), ("工事項目", 28), ("仕様", 18),
    ("単位", 6), ("数量", 10), ("単価", 14), ("金額", 14), ("備考", 16),
]


def _write_quote_header(ws: Any, quote: Any, project: Any) -> int:
    """見積書ヘッダ（ロゴ・タイトル・会社情報・工事名・合計額）を書き、次の行番号を返す。"""
    for i, w in enumerate(_QUOTE_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _insert_logo(ws, anchor="A1", width_px=110, height_px=55)

    ws.row_dimensions[1].height = 30
    title = ws.cell(row=1, column=4, value="見　積　書")
    title.font = Font(bold=True, size=20)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("D1:H1")

    _company_block(ws, 2, 5)

    ws.cell(row=2, column=1, value="見積番号").font = Font(size=9, bold=True)
    ws.cell(row=2, column=2, value=quote.quote_number or "")
    ws.cell(row=2, column=3, value="発行日").font = Font(size=9, bold=True)
    ws.cell(row=2, column=4, value=_fmt(quote.issue_date))

    ws.cell(row=3, column=1, value="有効期限").font = Font(size=9, bold=True)
    ws.cell(row=3, column=2, value=f"発行日より {quote.validity_days or 30} 日間")

    ws.row_dimensions[4].height = 4
    ws.cell(row=5, column=1, value=f"{project.client_name or ''} 御中")
    ws.cell(row=5, column=1).font = Font(bold=True, size=13)
    ws.merge_cells("A5:D5")

    ws.cell(row=6, column=1, value="下記の通り御見積申し上げます。")
    ws.cell(row=6, column=1).font = Font(size=9, italic=True)
    ws.merge_cells("A6:D6")

    ws.row_dimensions[7].height = 4
    ws.cell(row=8, column=1, value="工　事　名").font = Font(size=9, bold=True)
    ws.cell(row=8, column=2, value=project.project_name or "")
    ws.merge_cells("B8:H8")

    ws.cell(row=9, column=1, value="工事場所").font = Font(size=9, bold=True)
    ws.cell(row=9, column=2, value=quote.project_location_snapshot or project.project_location or "")
    ws.merge_cells("B9:H9")

    ws.cell(row=10, column=1, value="工　　期").font = Font(size=9, bold=True)
    period_text = ""
    if quote.period_start and quote.period_end:
        period_text = f"{_fmt(quote.period_start)} ～ {_fmt(quote.period_end)}"
    ws.cell(row=10, column=2, value=period_text)
    ws.merge_cells("B10:H10")

    ws.cell(row=11, column=1, value="支払条件").font = Font(size=9, bold=True)
    ws.cell(row=11, column=2, value=quote.payment_condition or "")
    ws.merge_cells("B11:H11")

    ws.row_dimensions[12].height = 4
    ws.row_dimensions[13].height = 28
    total_label = ws.cell(row=13, column=1, value="御　見　積　金　額")
    total_label.font = Font(bold=True, size=11)
    total_label.fill = _sub_fill()
    total_label.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A13:D13")

    total_cell = ws.cell(row=13, column=5, value=int(quote.total_amount or 0))
    total_cell.font = Font(bold=True, size=14)
    total_cell.number_format = "¥#,##0"
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    for cell in [ws.cell(row=13, column=c) for c in range(1, 9)]:
        cell.border = _thick_box()

    tax_note = ws.cell(row=13, column=6, value="（消費税10%込）")
    tax_note.font = Font(size=9, color="666666")
    tax_note.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("F13:H13")

    ws.row_dimensions[14].height = 4
    return 15  # 次の行


def _write_item_rows(ws: Any, start_row: int, items: list[Any]) -> int:
    """明細行を書き、次の行番号を返す。"""
    data_row = _apply_table_header(ws, start_row, _QUOTE_TABLE_COLS)
    filled = [i for i in items if i.item_name or i.amount]
    for idx, item in enumerate(filled, 1):
        r = data_row
        ws.row_dimensions[r].height = 16
        ws.cell(r, 1, idx).border = _thin()
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")

        c2 = ws.cell(r, 2, item.item_name or "")
        c2.border = _thin()
        c2.alignment = Alignment(vertical="center", wrap_text=True)

        c3 = ws.cell(r, 3, item.spec or "")
        c3.border = _thin()
        c3.alignment = Alignment(vertical="center", wrap_text=True)

        c4 = ws.cell(r, 4, item.unit or "")
        c4.border = _thin()
        c4.alignment = Alignment(horizontal="center", vertical="center")

        c5 = ws.cell(r, 5, float(item.quantity) if item.quantity else None)
        c5.border = _thin()
        c5.number_format = "#,##0.0##"
        c5.alignment = Alignment(horizontal="right", vertical="center")

        _money_cell(ws, r, 6, item.unit_price)
        _money_cell(ws, r, 7, item.amount or (
            (float(item.quantity or 0) * float(item.unit_price or 0))
            if item.quantity and item.unit_price else None
        ))

        c8 = ws.cell(r, 8, item.remarks or "")
        c8.border = _thin()
        c8.alignment = Alignment(vertical="center")

        data_row += 1
    return data_row


def _write_quote_summary(ws: Any, start_row: int, quote: Any) -> int:
    """小計・消費税・合計行を書き、次の行番号を返す。"""
    data_row = start_row + 1
    for label, val in [
        ("小　計（税抜）", quote.subtotal),
        ("消費税（10%）", quote.tax_amount),
        ("合　計（税込）", quote.total_amount),
    ]:
        ws.row_dimensions[data_row].height = 18
        lc = ws.cell(data_row, 6, label)
        lc.font = Font(bold=True, size=9)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border = _thin()
        ws.merge_cells(f"A{data_row}:F{data_row}")
        _money_cell(ws, data_row, 7, val)
        ws.cell(data_row, 7).font = Font(bold=True)
        ws.cell(data_row, 8).border = _thin()
        data_row += 1
    return data_row


def export_quote_excel(
    quote: Any,
    project: Any,
    items: list[Any],
    sections: list[Any] | None = None,
) -> bytes:
    """見積書をExcel(xlsx)として生成し、bytesで返す。

    大項目（sections）がある場合は 表紙 / 大項目集計 / 大項目別明細 の多シート構成。
    sections が空なら従来の1シート構成にフォールバック。
    """
    wb = Workbook()

    sections_sorted = sorted(sections or [], key=lambda s: s.row_no)

    if sections_sorted:
        # ── 多シート構成 ──────────────────────────────────────────────
        # Sheet1: 表紙（大項目の小計一覧付き）
        ws1 = wb.active
        ws1.title = "表紙"
        next_row = _write_quote_header(ws1, quote, project)

        # 大項目合計一覧を表紙に追加
        ws1.cell(next_row, 1, "大項目別集計").font = Font(bold=True, size=10)
        ws1.merge_cells(f"A{next_row}:H{next_row}")
        next_row += 1
        sec_hdr_row = next_row
        for col, (label, width) in enumerate([("記号", 6), ("大項目名", 34), ("小計", 14)], 1):
            c = ws1.cell(sec_hdr_row, col, label)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = _header_fill()
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin()
            ws1.column_dimensions[get_column_letter(col)].width = width
        next_row += 1
        for sec in sections_sorted:
            ws1.cell(next_row, 1, sec.section_letter).border = _thin()
            ws1.cell(next_row, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws1.cell(next_row, 1).font = Font(bold=True)
            ws1.cell(next_row, 2, sec.section_name).border = _thin()
            _money_cell(ws1, next_row, 3, sec.amount)
            ws1.merge_cells(f"D{next_row}:H{next_row}")
            next_row += 1

        # 合計行
        _write_quote_summary(ws1, next_row - 1, quote)

        ws1.print_area = f"A1:H{next_row + 8}"
        ws1.page_setup.orientation = "portrait"
        ws1.page_setup.fitToPage = True

        # Sheet2: 大項目集計
        ws2 = wb.create_sheet(title="大項目集計")
        for i, w in enumerate(_QUOTE_COL_WIDTHS, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.cell(1, 1, "大項目集計").font = Font(bold=True, size=14)
        ws2.merge_cells("A1:H1")
        ws2.cell(2, 1, f"見積番号：{quote.quote_number or ''}").font = Font(size=9)
        ws2.cell(2, 4, f"案件名：{project.project_name}").font = Font(size=9)
        r = 4
        r = _apply_table_header(ws2, r, [("記号", 6), ("大項目名", 40), ("小計（税抜）", 18), ("割合", 10)])
        subtotal_total = float(quote.subtotal or 0)
        for sec in sections_sorted:
            ws2.cell(r, 1, sec.section_letter).border = _thin()
            ws2.cell(r, 1).alignment = Alignment(horizontal="center")
            ws2.cell(r, 1).font = Font(bold=True)
            ws2.cell(r, 2, sec.section_name).border = _thin()
            _money_cell(ws2, r, 3, sec.amount)
            pct = (float(sec.amount or 0) / subtotal_total * 100) if subtotal_total else 0
            c4 = ws2.cell(r, 4, round(pct, 1))
            c4.number_format = "0.0%"
            c4.border = _thin()
            c4.alignment = Alignment(horizontal="right")
            r += 1
        r += 1
        for label, val in [
            ("小計（税抜）", quote.subtotal),
            ("消費税（10%）", quote.tax_amount),
            ("合計（税込）", quote.total_amount),
        ]:
            lc = ws2.cell(r, 2, label)
            lc.font = Font(bold=True, size=9)
            lc.border = _thin()
            _money_cell(ws2, r, 3, val)
            ws2.cell(r, 3).font = Font(bold=True)
            ws2.cell(r, 1).border = _thin()
            ws2.cell(r, 4).border = _thin()
            r += 1

        ws2.print_area = f"A1:D{r}"
        ws2.page_setup.fitToPage = True

        # Sheet3+: 大項目別明細
        item_by_section: dict[str, list[Any]] = {}
        unsectioned: list[Any] = []
        for item in items:
            sid = str(item.section_id) if item.section_id else None
            if sid:
                item_by_section.setdefault(sid, []).append(item)
            else:
                unsectioned.append(item)

        for sec in sections_sorted:
            sheet_name = f"{sec.section_letter}.{sec.section_name}"[:31]
            ws_sec = wb.create_sheet(title=sheet_name)
            for i, w in enumerate(_QUOTE_COL_WIDTHS, 1):
                ws_sec.column_dimensions[get_column_letter(i)].width = w

            ws_sec.cell(1, 1, f"{sec.section_letter}. {sec.section_name}").font = Font(bold=True, size=13)
            ws_sec.merge_cells("A1:H1")
            ws_sec.cell(2, 1, f"見積番号：{quote.quote_number or ''}  案件：{project.project_name}").font = Font(size=9, color="666666")
            ws_sec.merge_cells("A2:H2")
            ws_sec.row_dimensions[3].height = 4

            sec_items = sorted(item_by_section.get(str(sec.id), []), key=lambda x: x.row_no)
            dr = _write_item_rows(ws_sec, 4, sec_items)

            # 小計行
            dr += 1
            sec_subtotal = sum(float(i.amount or 0) for i in sec_items)
            lc = ws_sec.cell(dr, 6, "小　計")
            lc.font = Font(bold=True, size=9)
            lc.alignment = Alignment(horizontal="right")
            lc.border = _thin()
            ws_sec.merge_cells(f"A{dr}:F{dr}")
            _money_cell(ws_sec, dr, 7, sec_subtotal)
            ws_sec.cell(dr, 7).font = Font(bold=True)
            ws_sec.cell(dr, 8).border = _thin()

            ws_sec.print_area = f"A1:H{dr + 2}"
            ws_sec.page_setup.fitToPage = True

        if unsectioned:
            ws_other = wb.create_sheet(title="大項目未分類")
            for i, w in enumerate(_QUOTE_COL_WIDTHS, 1):
                ws_other.column_dimensions[get_column_letter(i)].width = w
            ws_other.cell(1, 1, "大項目未分類").font = Font(bold=True, size=13)
            ws_other.merge_cells("A1:H1")
            ws_other.row_dimensions[2].height = 4
            dr = _write_item_rows(ws_other, 3, unsectioned)
            ws_other.print_area = f"A1:H{dr + 2}"
            ws_other.page_setup.fitToPage = True

    else:
        # ── 従来の1シート構成（大項目なし） ──────────────────────────
        ws = wb.active
        ws.title = "見積書"
        next_row = _write_quote_header(ws, quote, project)
        data_row = _write_item_rows(ws, next_row, items)
        data_row = _write_quote_summary(ws, data_row, quote)

        data_row += 1
        if quote.conditions_text:
            ws.cell(data_row, 1, "工事条件等").font = Font(bold=True, size=9)
            data_row += 1
            for line in quote.conditions_text.split("\n"):
                c = ws.cell(data_row, 1, line)
                c.font = Font(size=9)
                ws.merge_cells(f"A{data_row}:H{data_row}")
                data_row += 1

        if quote.remarks:
            data_row += 1
            ws.cell(data_row, 1, "備考").font = Font(bold=True, size=9)
            data_row += 1
            ws.cell(data_row, 1, quote.remarks).font = Font(size=9)
            ws.merge_cells(f"A{data_row}:H{data_row}")

        ws.print_area = f"A1:H{data_row + 2}"
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 注文書 ────────────────────────────────────────────────────────────────────

def export_order_excel(order: Any, project: Any) -> bytes:
    """注文書をExcel(xlsx)として生成し、bytesで返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "注文書"

    col_widths = [4, 14, 20, 20, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── ロゴ ──
    _insert_logo(ws, anchor="A1", width_px=110, height_px=55)

    # ── タイトル ──
    ws.row_dimensions[1].height = 32
    title = ws.cell(row=1, column=3, value="注　文　書")
    title.font = Font(bold=True, size=20)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("C1:G1")

    # ── 注文番号・日付 ──
    ws.cell(row=2, column=1, value="注文番号").font = Font(size=9, bold=True)
    ws.cell(row=2, column=2, value=order.order_number or "")
    ws.cell(row=2, column=4, value="発行日").font = Font(size=9, bold=True)
    ws.cell(row=2, column=5, value=_fmt(order.issue_date))

    # ── 宛先 ──
    ws.row_dimensions[4].height = 4
    addr_label = ws.cell(row=5, column=1, value="宛　先")
    addr_label.font = Font(bold=True, size=9, color="555555")

    if order.client_address:
        ws.cell(row=6, column=1, value=order.client_address).font = Font(size=9)
        ws.merge_cells("A6:C6")

    name_row = 7 if order.client_address else 6
    ws.row_dimensions[name_row].height = 22
    cname = ws.cell(row=name_row, column=1, value=f"{order.client_company or ''}")
    cname.font = Font(bold=True, size=13)
    ws.merge_cells(f"A{name_row}:C{name_row}")

    person_row = name_row + 1
    if order.client_person:
        ws.cell(row=person_row, column=1, value=order.client_person).font = Font(size=10)
        ws.merge_cells(f"A{person_row}:C{person_row}")

    # ── 会社情報（右） ──
    _company_block(ws, 2, 4)

    # ── 工事件名 ──
    base_row = person_row + 2
    ws.cell(base_row, 1, "工　事　名").font = Font(size=9, bold=True)
    ws.cell(base_row, 2, project.project_name or "")
    ws.merge_cells(f"B{base_row}:G{base_row}")

    ws.cell(base_row + 1, 1, "工事場所").font = Font(size=9, bold=True)
    ws.cell(base_row + 1, 2, project.project_location or "")
    ws.merge_cells(f"B{base_row+1}:G{base_row+1}")

    # ── 金額 ──
    amt_row = base_row + 3
    ws.row_dimensions[amt_row].height = 26
    amt_label = ws.cell(amt_row, 1, "請負金額（税抜）")
    amt_label.font = Font(bold=True, size=11)
    amt_label.fill = _sub_fill()
    amt_label.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{amt_row}:C{amt_row}")

    amt_val = ws.cell(amt_row, 4, int(order.amount_excl_tax or 0))
    amt_val.font = Font(bold=True, size=13)
    amt_val.number_format = '¥#,##0'
    amt_val.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(f"D{amt_row}:G{amt_row}")

    for c in range(1, 8):
        ws.cell(amt_row, c).border = _thick_box()

    # 税額・合計
    r = amt_row + 1
    for label, val in [
        ("消費税（10%）", order.tax_amount),
        ("合計（税込）", order.total_amount),
    ]:
        ws.cell(r, 3, label).font = Font(size=9, bold=True)
        ws.cell(r, 3).alignment = Alignment(horizontal="right")
        ws.merge_cells(f"A{r}:C{r}")
        _money_cell(ws, r, 4, val)
        ws.cell(r, 4).font = Font(bold=True)
        ws.merge_cells(f"D{r}:G{r}")
        r += 1

    # 印紙税
    if order.stamp_tax:
        ws.cell(r, 3, "印紙税（自動算定）").font = Font(size=9, color="885500")
        ws.cell(r, 3).alignment = Alignment(horizontal="right")
        ws.merge_cells(f"A{r}:C{r}")
        _money_cell(ws, r, 4, order.stamp_tax)
        ws.cell(r, 4).font = Font(color="885500")
        ws.merge_cells(f"D{r}:G{r}")
        r += 1

    # ── 工期・支払条件 ──
    r += 1
    ws.cell(r, 1, "工　　期").font = Font(size=9, bold=True)
    period = ""
    if order.construction_period_start and order.construction_period_end:
        period = f"{_fmt(order.construction_period_start)} ～ {_fmt(order.construction_period_end)}"
    ws.cell(r, 2, period)
    ws.merge_cells(f"B{r}:G{r}")
    r += 1

    ws.cell(r, 1, "支払条件").font = Font(size=9, bold=True)
    ws.cell(r, 2, order.payment_condition or "")
    ws.merge_cells(f"B{r}:G{r}")
    r += 2

    # ── 約款 ──
    if order.terms_and_conditions:
        ws.cell(r, 1, "約款・特記事項").font = Font(bold=True, size=9)
        r += 1
        for line in order.terms_and_conditions.split("\n"):
            c = ws.cell(r, 1, line)
            c.font = Font(size=9)
            ws.merge_cells(f"A{r}:G{r}")
            r += 1

    ws.print_area = f"A1:G{r + 2}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 注文請書 ──────────────────────────────────────────────────────────────────

def export_acknowledgment_excel(ack: Any, project: Any) -> bytes:
    """注文請書をExcel(xlsx)として生成し、bytesで返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "注文請書"

    col_widths = [4, 14, 20, 20, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── ロゴ ──
    _insert_logo(ws, anchor="A1", width_px=110, height_px=55)

    # ── タイトル ──
    ws.row_dimensions[1].height = 32
    title = ws.cell(row=1, column=3, value="注 文 請 書")
    title.font = Font(bold=True, size=20)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("C1:G1")

    # ── 注文請書番号・日付 ──
    ws.cell(row=2, column=1, value="請書番号").font = Font(size=9, bold=True)
    ws.cell(row=2, column=2, value=ack.acknowledgment_number or "")
    ws.cell(row=2, column=4, value="発行日").font = Font(size=9, bold=True)
    ws.cell(row=2, column=5, value=_fmt(ack.issue_date))

    # ── 会社情報（受注者＝クラップ、右側に表示） ──
    _company_block(ws, 2, 4)

    # ── 受注者情報（発行者挨拶文） ──
    ws.row_dimensions[4].height = 4
    ws.cell(row=5, column=1, value="下記の通り確かに注文をお受け致しました。")
    ws.cell(row=5, column=1).font = Font(size=9, italic=True)
    ws.merge_cells("A5:G5")

    # ── 宛先（発注者） ──
    ws.row_dimensions[6].height = 4
    addr_label = ws.cell(row=7, column=1, value="宛　先（発注者）")
    addr_label.font = Font(bold=True, size=9, color="555555")

    if ack.client_address:
        ws.cell(row=8, column=1, value=ack.client_address).font = Font(size=9)
        ws.merge_cells("A8:C8")

    name_row = 9 if ack.client_address else 8
    ws.row_dimensions[name_row].height = 22
    cname = ws.cell(row=name_row, column=1, value=f"{ack.client_company or ''}")
    cname.font = Font(bold=True, size=13)
    ws.merge_cells(f"A{name_row}:C{name_row}")

    person_row = name_row + 1
    if ack.client_person:
        ws.cell(row=person_row, column=1, value=ack.client_person).font = Font(size=10)
        ws.merge_cells(f"A{person_row}:C{person_row}")

    # ── 工事件名 ──
    base_row = person_row + 2
    ws.cell(base_row, 1, "工　事　名").font = Font(size=9, bold=True)
    ws.cell(base_row, 2, project.project_name or "")
    ws.merge_cells(f"B{base_row}:G{base_row}")

    ws.cell(base_row + 1, 1, "工事場所").font = Font(size=9, bold=True)
    ws.cell(base_row + 1, 2, project.project_location or "")
    ws.merge_cells(f"B{base_row+1}:G{base_row+1}")

    # ── 金額 ──
    amt_row = base_row + 3
    ws.row_dimensions[amt_row].height = 26
    amt_label = ws.cell(amt_row, 1, "請負金額（税抜）")
    amt_label.font = Font(bold=True, size=11)
    amt_label.fill = _sub_fill()
    amt_label.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{amt_row}:C{amt_row}")

    amt_val = ws.cell(amt_row, 4, int(ack.amount_excl_tax or 0))
    amt_val.font = Font(bold=True, size=13)
    amt_val.number_format = '¥#,##0'
    amt_val.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(f"D{amt_row}:G{amt_row}")

    for c in range(1, 8):
        ws.cell(amt_row, c).border = _thick_box()

    r = amt_row + 1
    for label, val in [
        ("消費税（10%）", ack.tax_amount),
        ("合計（税込）", ack.total_amount),
    ]:
        ws.cell(r, 3, label).font = Font(size=9, bold=True)
        ws.cell(r, 3).alignment = Alignment(horizontal="right")
        ws.merge_cells(f"A{r}:C{r}")
        _money_cell(ws, r, 4, val)
        ws.cell(r, 4).font = Font(bold=True)
        ws.merge_cells(f"D{r}:G{r}")
        r += 1

    if ack.stamp_tax:
        ws.cell(r, 3, "印紙税（自動算定）").font = Font(size=9, color="885500")
        ws.cell(r, 3).alignment = Alignment(horizontal="right")
        ws.merge_cells(f"A{r}:C{r}")
        _money_cell(ws, r, 4, ack.stamp_tax)
        ws.cell(r, 4).font = Font(color="885500")
        ws.merge_cells(f"D{r}:G{r}")
        r += 1

    # ── 工期・支払条件 ──
    r += 1
    ws.cell(r, 1, "工　　期").font = Font(size=9, bold=True)
    period = ""
    if ack.construction_period_start and ack.construction_period_end:
        period = f"{_fmt(ack.construction_period_start)} ～ {_fmt(ack.construction_period_end)}"
    ws.cell(r, 2, period)
    ws.merge_cells(f"B{r}:G{r}")
    r += 1

    ws.cell(r, 1, "支払条件").font = Font(size=9, bold=True)
    ws.cell(r, 2, ack.payment_condition or "")
    ws.merge_cells(f"B{r}:G{r}")
    r += 2

    # ── 約款 ──
    if ack.terms_and_conditions:
        ws.cell(r, 1, "約款・特記事項").font = Font(bold=True, size=9)
        r += 1
        for line in ack.terms_and_conditions.split("\n"):
            c = ws.cell(r, 1, line)
            c.font = Font(size=9)
            ws.merge_cells(f"A{r}:G{r}")
            r += 1

    ws.print_area = f"A1:G{r + 2}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 請求書 ────────────────────────────────────────────────────────────────────

def export_invoice_excel(invoice: Any, project: Any) -> bytes:
    """請求書をExcel(xlsx)として生成し、bytesで返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "請求書"

    col_widths = [4, 18, 18, 18, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── ロゴ ──
    _insert_logo(ws, anchor="A1", width_px=110, height_px=55)

    # ── タイトル ──
    ws.row_dimensions[1].height = 32
    title = ws.cell(row=1, column=3, value="御　請　求　書")
    title.font = Font(bold=True, size=20)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("C1:F1")

    # ── 会社情報（右）— マージ操作より前に配置 ──
    _company_block(ws, 2, 4)

    # ── 請求番号・発行日 ──
    ws.cell(2, 1, "請求番号").font = Font(size=9, bold=True)
    ws.cell(2, 2, invoice.invoice_number or "")

    # ── 宛先 ──
    ws.row_dimensions[4].height = 4
    ws.cell(5, 1, f"{project.client_name or ''} 御中")
    ws.cell(5, 1).font = Font(bold=True, size=13)
    ws.merge_cells("A5:C5")

    ws.cell(6, 1, "下記の通り御請求申し上げます。")
    ws.cell(6, 1).font = Font(size=9, italic=True)
    ws.merge_cells("A6:C6")

    # ── 工事件名 ──
    ws.row_dimensions[7].height = 4
    ws.cell(8, 1, "工　事　名").font = Font(size=9, bold=True)
    ws.cell(8, 2, project.project_name or "")
    ws.merge_cells("B8:F8")

    # ── 入金管理テーブル ──
    ws.row_dimensions[10].height = 4
    headers = [("項　目", 14), ("金　額", 14), ("", 18), ("", 14), ("", 14), ("", 14)]
    pay_row = _apply_table_header(ws, 11, headers)

    pay_items = [
        ("前月御請求額", invoice.previous_balance),
        ("御入金額", invoice.received_amount),
        ("差引残高", invoice.outstanding_balance),
    ]
    for label, val in pay_items:
        ws.row_dimensions[pay_row].height = 18
        lc = ws.cell(pay_row, 1, label)
        lc.font = Font(size=10)
        lc.border = _thin()
        lc.alignment = Alignment(vertical="center")
        ws.merge_cells(f"A{pay_row}:B{pay_row}")
        _money_cell(ws, pay_row, 3, val)
        ws.merge_cells(f"C{pay_row}:F{pay_row}")
        for c in range(1, 7):
            ws.cell(pay_row, c).border = _thin()
        pay_row += 1

    # ── 当月請求 ──
    pay_row += 1
    ws.row_dimensions[pay_row].height = 26
    cur_label = ws.cell(pay_row, 1, "今月御請求額（税抜）")
    cur_label.font = Font(bold=True, size=11)
    cur_label.fill = _sub_fill()
    cur_label.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{pay_row}:B{pay_row}")

    cur_val = ws.cell(pay_row, 3, int(invoice.current_purchase or 0))
    cur_val.font = Font(bold=True, size=13)
    cur_val.number_format = '¥#,##0'
    cur_val.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(f"C{pay_row}:F{pay_row}")
    for c in range(1, 7):
        ws.cell(pay_row, c).border = _thick_box()
    pay_row += 1

    for label, val in [("消費税（10%）", invoice.tax_amount), ("合計（税込）", invoice.total_amount)]:
        ws.row_dimensions[pay_row].height = 18
        lc = ws.cell(pay_row, 2, label)
        lc.font = Font(bold=True, size=9)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        ws.merge_cells(f"A{pay_row}:B{pay_row}")
        _money_cell(ws, pay_row, 3, val)
        ws.cell(pay_row, 3).font = Font(bold=True)
        ws.merge_cells(f"C{pay_row}:F{pay_row}")
        for c in range(1, 7):
            ws.cell(pay_row, c).border = _thin()
        pay_row += 1

    # ── 振込先 ──
    pay_row += 2
    ws.cell(pay_row, 1, "【お振込先】").font = Font(bold=True, size=9)
    ws.merge_cells(f"A{pay_row}:F{pay_row}")
    pay_row += 1
    ws.cell(pay_row, 1, COMPANY_BANK).font = Font(size=9)
    ws.merge_cells(f"A{pay_row}:F{pay_row}")

    ws.print_area = f"A1:F{pay_row + 2}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_project_all_excel(
    project: Any,
    qcds_rows: list[Any],
    quotes: list[Any],
    orders: list[Any],
    invoices: list[Any],
) -> bytes:
    """案件の全データ（案件情報・QCDS・見積・注文・請求）を1つのExcelブックに出力する。"""
    wb = Workbook()

    # ── シート1: 案件情報 ────────────────────────────────────────────────────
    ws_proj = wb.active
    ws_proj.title = "案件情報"
    ws_proj.column_dimensions["A"].width = 20
    ws_proj.column_dimensions["B"].width = 40

    def _hdr(r: int, label: str) -> None:
        c = ws_proj.cell(r, 1, label)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = _header_fill()
        c.border = _thin()
        ws_proj.cell(r, 2).border = _thin()
        ws_proj.row_dimensions[r].height = 16

    def _row(r: int, label: str, value: Any) -> None:
        lc = ws_proj.cell(r, 1, label)
        lc.font = Font(size=9, bold=True)
        lc.fill = _sub_fill()
        lc.border = _thin()
        vc = ws_proj.cell(r, 2, _fmt(value))
        vc.font = Font(size=9)
        vc.border = _thin()
        ws_proj.row_dimensions[r].height = 14

    _hdr(1, "案件情報")
    ws_proj.merge_cells("A1:B1")
    r = 2
    fields = [
        ("工事番号", getattr(project, "project_number", None)),
        ("工事名", getattr(project, "project_name", None)),
        ("工事場所", getattr(project, "project_location", None)),
        ("発注者", getattr(project, "client_name", None)),
        ("元発注者", getattr(project, "original_client_name", None)),
        ("工事価格", getattr(project, "project_price", None)),
        ("ステータス", str(project.status) if project.status else None),
        ("工期(契約)開始", getattr(project, "period_contract_start", None)),
        ("工期(契約)終了", getattr(project, "period_contract_end", None)),
        ("支払条件", getattr(project, "payment_condition", None)),
    ]
    for label, value in fields:
        _row(r, label, value)
        r += 1

    # ── シート2: QCDS ────────────────────────────────────────────────────────
    if qcds_rows:
        ws_qcds = wb.create_sheet("QCDS")
        ws_qcds.column_dimensions["A"].width = 5
        cols = [
            ("行番", 5), ("業者名", 20), ("工種", 15),
            ("実行予算", 14), ("取決金額", 14), ("精算見通", 14),
        ]
        next_r = _apply_table_header(ws_qcds, 1, cols)
        for q in sorted(qcds_rows, key=lambda x: x.row_no):
            ws_qcds.cell(next_r, 1, q.row_no).border = _thin()
            ws_qcds.cell(next_r, 2, q.vendor_name_snapshot or "").border = _thin()
            ws_qcds.cell(next_r, 3, q.work_type or "").border = _thin()
            _money_cell(ws_qcds, next_r, 4, q.budget_amount)
            _money_cell(ws_qcds, next_r, 5, q.agreed_amount)
            _money_cell(ws_qcds, next_r, 6, q.settlement_amount)
            for c in range(1, 7):
                ws_qcds.cell(next_r, c).font = Font(size=9)
            next_r += 1

    # ── シート3〜: 見積書 ──────────────────────────────────────────────────
    for i, quote in enumerate(quotes, 1):
        sheet_name = f"見積書{'_' + str(i) if len(quotes) > 1 else ''}"[:31]
        ws_q = wb.create_sheet(sheet_name)
        items = sorted(quote.items, key=lambda x: x.row_no) if hasattr(quote, "items") else []
        # 見積書の要約情報
        ws_q.column_dimensions["A"].width = 20
        ws_q.column_dimensions["B"].width = 40
        summary = [
            ("見積番号", quote.quote_number),
            ("発行日", quote.issue_date),
            ("ステータス", str(quote.status) if quote.status else None),
            ("消費税抜合計", getattr(quote, "subtotal", None)),
            ("消費税", getattr(quote, "tax_amount", None)),
            ("合計(税込)", getattr(quote, "total_amount", None)),
        ]
        r2 = 1
        for label, value in summary:
            ws_q.cell(r2, 1, label).font = Font(bold=True, size=9)
            ws_q.cell(r2, 2, _fmt(value)).font = Font(size=9)
            r2 += 1
        # 内訳
        r2 += 1
        item_cols = [("行番", 5), ("品目", 28), ("規格", 18), ("単位", 6), ("数量", 8), ("単価", 12), ("金額", 14), ("備考", 20)]
        r2 = _apply_table_header(ws_q, r2, item_cols)
        for item in items:
            ws_q.cell(r2, 1, item.row_no).border = _thin()
            ws_q.cell(r2, 2, item.item_name or "").border = _thin()
            ws_q.cell(r2, 3, item.spec or "").border = _thin()
            ws_q.cell(r2, 4, item.unit or "").border = _thin()
            ws_q.cell(r2, 5, item.quantity).border = _thin()
            _money_cell(ws_q, r2, 6, item.unit_price)
            _money_cell(ws_q, r2, 7, item.amount)
            ws_q.cell(r2, 8, item.remarks or "").border = _thin()
            for c in range(1, 9):
                ws_q.cell(r2, c).font = Font(size=9)
            r2 += 1

    # ── シート: 注文書 ─────────────────────────────────────────────────────
    if orders:
        ws_ord = wb.create_sheet("注文書一覧")
        ws_ord.column_dimensions["A"].width = 15
        ord_cols = [
            ("注文番号", 15), ("発行日", 12), ("発注先", 20),
            ("税抜金額", 14), ("ステータス", 12),
        ]
        next_r = _apply_table_header(ws_ord, 1, ord_cols)
        for o in orders:
            ws_ord.cell(next_r, 1, o.order_number or "").border = _thin()
            ws_ord.cell(next_r, 2, _fmt(o.issue_date)).border = _thin()
            ws_ord.cell(next_r, 3, o.client_company or "").border = _thin()
            _money_cell(ws_ord, next_r, 4, o.amount_excl_tax)
            ws_ord.cell(next_r, 5, str(o.status) if o.status else "").border = _thin()
            for c in range(1, 6):
                ws_ord.cell(next_r, c).font = Font(size=9)
            next_r += 1

    # ── シート: 請求書 ─────────────────────────────────────────────────────
    if invoices:
        ws_inv = wb.create_sheet("請求書一覧")
        inv_cols = [
            ("請求番号", 15), ("発行日", 12), ("入金期限", 12),
            ("税抜合計", 14), ("消費税", 12), ("税込合計", 14), ("ステータス", 12),
        ]
        next_r = _apply_table_header(ws_inv, 1, inv_cols)
        for inv in invoices:
            ws_inv.cell(next_r, 1, inv.invoice_number or "").border = _thin()
            ws_inv.cell(next_r, 2, _fmt(inv.issue_date)).border = _thin()
            ws_inv.cell(next_r, 3, _fmt(getattr(inv, "payment_due_date", None))).border = _thin()
            _money_cell(ws_inv, next_r, 4, getattr(inv, "current_purchase", None))
            _money_cell(ws_inv, next_r, 5, getattr(inv, "tax_amount", None))
            _money_cell(ws_inv, next_r, 6, getattr(inv, "total_amount", None))
            ws_inv.cell(next_r, 7, str(inv.status) if inv.status else "").border = _thin()
            for c in range(1, 8):
                ws_inv.cell(next_r, c).font = Font(size=9)
            next_r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_qcds_excel(project: Any, qcds: Any) -> bytes:
    """QCDS 原価算定表を Excel で出力する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "QCDS原価算定表"
    for col, w in zip("ABCDEFG", [5, 22, 16, 10, 16, 16, 16]):
        ws.column_dimensions[col].width = w

    title_cell = ws.cell(1, 1, "QCDS 原価算定表")
    title_cell.font = Font(bold=True, size=12)
    ws.merge_cells("A1:G1")
    ws.cell(2, 1, f"工事番号: {getattr(project, 'project_number', '')}  工事名: {getattr(project, 'project_name', '')}").font = Font(size=9, color="666666")
    ws.merge_cells("A2:G2")

    dw_cols = [("No",5),("支払先",22),("工種",16),("区分",10),("実行予算",16),("取決金額",16),("精算見通",16)]
    row = _apply_table_header(ws, 4, dw_cols)

    direct_works = sorted(getattr(qcds, "direct_works", []) or [], key=lambda x: getattr(x, "row_no", 0))
    for i, w in enumerate(direct_works, 1):
        ws.cell(row, 1, i).border = _thin()
        ws.cell(row, 2, getattr(w, "vendor_name_snapshot", "") or "").border = _thin()
        ws.cell(row, 3, getattr(w, "work_type", "") or "").border = _thin()
        ws.cell(row, 4, str(getattr(w, "category", "") or "")).border = _thin()
        _money_cell(ws, row, 5, getattr(w, "budget_amount", None))
        _money_cell(ws, row, 6, getattr(w, "agreed_amount", None))
        _money_cell(ws, row, 7, getattr(w, "settlement_amount", None))
        for c in range(1, 8):
            ws.cell(row, c).font = Font(size=9)
        row += 1

    total_budget = sum(float(getattr(w, "budget_amount", 0) or 0) for w in direct_works)
    total_agreed = sum(float(getattr(w, "agreed_amount", 0) or 0) for w in direct_works)
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row, 1, "合計").font = Font(bold=True, size=9)
    _money_cell(ws, row, 5, total_budget)
    _money_cell(ws, row, 6, total_agreed)
    ws.cell(row, 5).font = Font(bold=True, size=9)
    ws.cell(row, 6).font = Font(bold=True, size=9)

    calc = getattr(qcds, "calc", None)
    if calc:
        row += 2
        hc = ws.cell(row, 1, "工事割出サマリー")
        hc.fill = _header_fill()
        hc.font = Font(bold=True, size=10, color="FFFFFF")
        ws.merge_cells(f"A{row}:G{row}")
        row += 1
        for label, attr in [
            ("直接工事費（実行予算）", "direct_cost_budget"),
            ("直接工事費（取決見通）", "direct_cost_agreed"),
            ("現場経費合計", "site_overhead_total"),
            ("原価合計", "construction_cost_total"),
            ("一般管理費", "general_admin_cost"),
            ("営業利益", "operating_profit"),
        ]:
            lc = ws.cell(row, 1, label)
            lc.font = Font(size=9, bold=True); lc.fill = _sub_fill(); lc.border = _thin()
            ws.merge_cells(f"A{row}:D{row}")
            _money_cell(ws, row, 5, getattr(calc, attr, None))
            ws.cell(row, 5).font = Font(size=9)
            row += 1

    buf2 = io.BytesIO()
    wb.save(buf2)
    return buf2.getvalue()
